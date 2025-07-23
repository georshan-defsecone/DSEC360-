import os
import subprocess
import re
import shutil
import zipfile
from io import BytesIO
from .Configuration_Audit.Database.MARIA import connection_maria
from . import remote
from .Configuration_Audit.Database.ORACLE import generate_sql
import zipfile
from django.http import HttpResponse, JsonResponse
from django.views import View
from django.conf import settings
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from django.http import JsonResponse, Http404
import json
import pandas as pd
import os
from openpyxl.styles import Font, PatternFill,Alignment
from openpyxl.utils import get_column_letter
from .Compromise_Assesment.Windows.make import generate_powershell_script

from .Configuration_Audit.Windows.generate_PowerShell import generate_script
from .Configuration_Audit.Windows.validate import validate_compliance
from .Configuration_Audit.Windows.Windows_Remote_exec.wmi import run_remote_audit, cleanup_remote_files
from .remote import mariadb_connection,linux_connection,oracle_connection,mssql_connection

def database_config_audit(data):
    print("[*] Entered database_config_audit()")
    scan_data = data.get("scan_data", {})

    # Normalize compliance and standard names
    compliance_name = (scan_data.get("complianceCategory") or "").strip().lower()
    standard = (scan_data.get("complianceSecurityStandard") or "").strip().lower()

    normalized_compliance = re.sub(r'[\W_]+', '', compliance_name)
    normalized_standard = re.sub(r'[\W_]+', '', standard)
    print(f"[DEBUG] normalized_compliance: {normalized_compliance}")
    unchecked_items = scan_data.get("uncheckedComplianceItems", [])
    project_name = (data.get("project_name") or "unknown_project").strip().replace(" ", "_")
    scan_name = (data.get("scan_name") or "unknown_scan").strip().replace(" ", "_")
    safe_compliance_name = normalized_compliance or "unknown_compliance"
    safe_standard = normalized_standard or "unknown_standard"
    safe_project_name = project_name
    safe_scan_name = scan_name
    base_dir = os.path.dirname(os.path.abspath(__file__))
    project_folder = os.path.join(
                                base_dir,
                                "Projects",
                                project_name,
                                scan_name
                            )

    base_dynamic_filename = f"{safe_compliance_name}_{safe_standard}_{safe_project_name}_{safe_scan_name}"
    dynamic_filename_extension = ".csv"
    dynamic_filename = f"{base_dynamic_filename}{dynamic_filename_extension}"
    result_csv = os.path.join(project_folder, dynamic_filename)
    version = 1
    while os.path.exists(result_csv):
            version += 1
            dynamic_filename = f"{base_dynamic_filename}_v{version}{dynamic_filename_extension}"
            result_csv = os.path.join(project_folder, dynamic_filename)

    if normalized_compliance == "oracle_12c" or normalized_compliance == "oracle12c":
        try:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            oracle_dir = os.path.join(base_dir,"Configuration_Audit","Database","ORACLE","CIS","Queries")
            csv_name = "oracle_12c_cis.csv"
            csv_path = os.path.join(oracle_dir, csv_name)
            
            project_name = (data.get("project_name") or "unknown_project").strip().replace(" ", "_")
            scan_name = (data.get("scan_name") or "unknown_scan").strip().replace(" ", "_")

            project_folder = os.path.join(
                                base_dir,
                                "Projects",
                                project_name,
                                scan_name
                            )


            # Create the nested directory structure if it doesn't exist
            os.makedirs(project_folder, exist_ok=True)

            
            # Normalize components for filename
            safe_compliance_name = normalized_compliance or "unknown_compliance"
            safe_standard = normalized_standard or "unknown_standard"
            safe_project_name = project_name
            safe_scan_name = scan_name

            base_dynamic_filename = f"{safe_compliance_name}_{safe_standard}_{safe_project_name}_{safe_scan_name}"
            dynamic_filename_extension = ".csv"
            dynamic_filename = f"{base_dynamic_filename}{dynamic_filename_extension}"
            result_csv = os.path.join(project_folder, dynamic_filename)
            sql_output = os.path.join(project_folder, f"{base_dynamic_filename}.sql")
            version = 1
            while os.path.exists(result_csv):
                  version += 1
                  dynamic_filename = f"{base_dynamic_filename}_v{version}{dynamic_filename_extension}"
                  result_csv = os.path.join(project_folder, dynamic_filename)
        # --- MODIFICATION END ---

            json_output = os.path.join(project_folder, "output.json")

            print(json_output)

            if not os.path.exists(csv_path):
                print(f"[!] CSV input file not found: {csv_path}")
                return None,None

            # Step 1: Generate the SQL script
            queries = generate_sql.extract_db_queries(csv_path, unchecked_items)
            if not queries:
                print("[!] No queries extracted from CSV.")
                return None,None

            generate_sql.write_queries_to_file(queries, sql_output, unchecked_items)

            print(f"[+] SQL script generated at: {sql_output}")

            # Step 2: Choose execution method based on auditMethod
            audit_method = (scan_data.get("auditMethod") or "").strip().lower()
            print(audit_method)

            if audit_method == "remoteaccess":
                print("[*] Using remote access method.")
                remote.oracle_connection(sql_output, scan_data, json_output, result_csv)
                print("returning ",result_csv,json_output)
                convert_csv_to_excel(result_csv)
                

                return result_csv,json_output
            elif audit_method == "agent":
                print("[*] Using agent method.")
                script_path= download_script(sql_output)
                print(f"[+] Script downloaded to: {script_path}")
                return script_path, json_output
            else:
                print(f"[-] Unsupported audit method: {audit_method}")
                return None,None

        except Exception as e:
            print(f"[!] Exception running Oracle audit: {e}")
            return None,None
    elif normalized_compliance=="mariadb106" or normalized_compliance=="mariadb1011":
        try:
            print(f"[DEBUG] Running MariaDB audit for compliance: {normalized_compliance}")
            base_dir = os.path.dirname(os.path.abspath(__file__))
            Maria_dir = os.path.join(base_dir,"Configuration_Audit","Database","MARIA")

            excluded_audit = scan_data.get("uncheckedComplianceItems") or []
            user_name = scan_data.get("username")
            password_name = scan_data.get("password")
            host_name = scan_data.get("target")
            port_number = scan_data.get("port") or 3306
            domain_name = scan_data.get("domain") or ""
            db_access_method = scan_data.get("auditMethod")
            database_name=scan_data.get("database") or 'mysql'
            project_name = (data.get("project_name") or "unknown_project").strip().replace(" ", "_")
            scan_name = (data.get("scan_name") or "unknown_scan").strip().replace(" ", "_")
            print(f"[DEBUG] excluded_audit: {excluded_audit}, user_name: {user_name},password:{password_name}, host_name: {host_name}, port_number: {port_number}, domain_name: {domain_name}, db_access_method: {db_access_method}")
            # Check which version of MariaDB to run

            if normalized_compliance == "mariadb106":
                print("testing mariadb connection for 10.6")
                input_csv_path = os.path.join(Maria_dir, "CIS", "Queries", "MariaDB_10_6_query.csv")
                sql_commands = os.path.join(Maria_dir, "CIS", "MariaDB_10_6_cis_query.sql")
                linux_file = os.path.join(Maria_dir, "CIS", "MariaDB_10_6_linux_commands.sh")
                try:
                  remote.mariadb_connection(result_csv,excluded_audit, user_name, password_name, host_name, port_number,database_name, domain_name, db_access_method, input_csv_path, sql_commands, linux_file,normalized_compliance)
                except SystemExit as e:
                  print(f"Validation caused exit with code: {e.code}")
                
                if db_access_method == "agent":
                    path_for_sql = os.path.join(Maria_dir, "CIS", "MariaDB_10_6_cis_query.sql")
                    #path_for_linux = os.path.join(Maria_dir, "CIS", "linux_commands.sh")
                    path_for_script=download_script(path_for_sql)
                    return path_for_script,None
                if db_access_method == "remoteAccess":
                    path_for_json = os.path.join(Maria_dir, "CIS", "mariaDB_10_6_query_result.json")
                    path_for_csv= result_csv
                    return path_for_csv,path_for_json
                    
            if normalized_compliance == "mariadb1011":
                input_csv_path = os.path.join(Maria_dir, "CIS", "Queries", "MariaDB_10_11_query.csv")
                sql_commands = os.path.join(Maria_dir, "CIS", "MariaDB_10_11_cis_query.sql")
                linux_file = os.path.join(Maria_dir, "CIS", "MariaDB_10_11_linux_commands.sh")
                remote.mariadb_connection(result_csv,excluded_audit, user_name, password_name, host_name, port_number,database_name, domain_name, db_access_method, input_csv_path, sql_commands, linux_file,normalized_compliance)
                if db_access_method == "agent":
                    path_for_sql = os.path.join(Maria_dir, "CIS", "MariaDB_10_11_cis_query.sql")
                    path_for_script= download_script(path_for_sql)
                    #path_for_linux = os.path.join(Maria_dir, "CIS_standard", "linux_commands.sh")
                    return path_for_script,None
                if db_access_method=="remoteAccess":
                    path_for_json = os.path.join(Maria_dir, "CIS", "mariaDB_10_11_query_result.json")
                    path_for_csv= result_csv
                    return path_for_csv,path_for_json
        except Exception as e:
            print(f"[!] Exception running Oracle audit: {e}")
            return None

        
    # Check for MSSQL compliance
    elif normalized_compliance == "microsoftsqlserver2019" or normalized_compliance == "microsoftsqlserver2017" or normalized_compliance == "microsoftsqlserver2016" or normalized_compliance == "microsoftsqlserver2022":
        try:
            print("[*] Running MSSQL audit")
            #assining the initial file paths
            base_dir = os.path.dirname(os.path.abspath(__file__))
            mssql_dir = os.path.join(base_dir, "Configuration_Audit", "Database", "MSSQL")
            #getting the values from the scan_data
            excluded_audit = scan_data.get("uncheckedComplianceItems") or []
            user_name = scan_data.get("username")
            password_name = scan_data.get("password")
            host_name = scan_data.get("target")
            port_number = scan_data.get("port") or 1433
            domain_name = scan_data.get("domain") or ""
            db_access_method = scan_data.get("auditMethod")
            database_name=scan_data.get("database") or 'master'

            if normalized_compliance == "microsoftsqlserver2019":
                print("testing mssql connection for 2019")
                remote.mssql_connection(result_csv,excluded_audit, user_name, password_name, host_name, port_number, database_name, domain_name, db_access_method, normalized_compliance)
                if db_access_method == "agent":
                    path_for_sql=os.path.join(mssql_dir,"CIS","microsoft_sql_server_2019_cis_query.sql")
                    path_for_script= download_script(path_for_sql)
                    return path_for_script,None
                if db_access_method == "remoteAccess":
                    path_for_json=os.path.join(mssql_dir,"CIS","microsoft_sql_server_2019_query_result.json")
                    path_for_csv=result_csv
                    return path_for_csv,path_for_json
            elif normalized_compliance == "microsoftsqlserver2017":
                remote.mssql_connection(result_csv,excluded_audit, user_name, password_name, host_name, port_number, database_name, domain_name, db_access_method, normalized_compliance)
                if db_access_method == "agent":
                    path_for_sql=os.path.join(mssql_dir,"CIS","microsoft_sql_server_2017_cis_query.sql")
                    path_for_script= download_script(path_for_sql)
                    return path_for_script,None
                if db_access_method == "remoteAccess":
                    path_for_json=os.path.join(mssql_dir,"CIS","microsoft_sql_server_2017_query_result.json")
                    path_for_csv=result_csv
                    return path_for_csv,path_for_json

            elif normalized_compliance == "microsoftsqlserver2016":
                remote.mssql_connection(result_csv,excluded_audit, user_name, password_name, host_name, port_number, database_name, domain_name, db_access_method, normalized_compliance)
                if db_access_method == "agent":
                    path_for_sql=os.path.join(mssql_dir,"CIS","microsoft_sql_server_2016_cis_query.sql")
                    path_for_script= download_script(path_for_sql)
                    return path_for_script,None
                if db_access_method == "remoteAccess":
                    path_for_json=os.path.join(mssql_dir,"CIS","microsoft_sql_server_2016_query_result.json")
                    path_for_csv=result_csv
                    return path_for_csv,path_for_json

            elif normalized_compliance == "microsoftsqlserver2022":
                remote.mssql_connection(result_csv,excluded_audit, user_name, password_name, host_name, port_number, database_name, domain_name, db_access_method, normalized_compliance)
                if db_access_method == "agent":
                    path_for_sql=os.path.join(mssql_dir,"CIS","microsoft_sql_server_2022_cis_query.sql")
                    path_for_script= download_script(path_for_sql)
                    return path_for_script,None
                if db_access_method == "remoteAccess":
                    path_for_json=os.path.join(mssql_dir,"CIS","microsoft_sql_server_2022_query_result.json")
                    path_for_csv=result_csv
                    return path_for_csv,path_for_json

        except Exception as e:
            print(f"[!] Exception running MSSQL audit: {e}")
            return None
    else:
        print(f"[-] Unsupported compliance type: {normalized_compliance}")
        return None,None

# In views.py

def linux_config_audit(data):
    print("[*] Entered linux_config_audit()")
    from .Configuration_Audit.Linux.generate import ubuntu
    import os
    scan_data = data.get("scan_data", {})

    # Get standard, version, and method
    standard = scan_data.get("complianceSecurityStandard", "CIS").strip()
    version = scan_data.get("complianceCategory", "").strip().replace(" ", "_")
    method = scan_data.get("auditMethod", "remote").strip().lower()
    excluded = scan_data.get("uncheckedComplianceItems", [])

    # Get project and scan names for dynamic paths
    project_name = (data.get("project_name") or "unknown_project").strip().replace(" ", "_")
    scan_name = (data.get("scan_name") or "unknown_scan").strip().replace(" ", "_")
    
    # Define base directory and create dynamic project/scan folder
    base_dir = os.path.dirname(os.path.abspath(__file__))
    project_folder = os.path.join(base_dir, "Projects", project_name, scan_name)
    os.makedirs(project_folder, exist_ok=True)

    # --- START of MODIFICATIONS ---

    # Create base names for the files
    base_script_name = f"combined_{standard}_{version}"
    base_json_name = f"results_{standard}_{version}"
    base_csv_name = f"results_{standard}_{version}"

    # Generate initial, non-versioned file paths
    script_path = os.path.join(project_folder, f"{base_script_name}.sh")
    json_path = os.path.join(project_folder, f"{base_json_name}.json")
    csv_path = os.path.join(project_folder, f"{base_csv_name}.csv")

    # Add versioning logic for rescans
    scan_version = 1
    # Check if the primary result file (CSV) already exists
    while os.path.exists(csv_path):
        scan_version += 1
        # Create new versioned filenames for all related files
        versioned_script_name = f"{base_script_name}_v{scan_version}.sh"
        versioned_json_name = f"{base_json_name}_v{scan_version}.json"
        versioned_csv_name = f"{base_csv_name}_v{scan_version}.csv"
        
        # Update the full paths with the new versioned names
        script_path = os.path.join(project_folder, versioned_script_name)
        json_path = os.path.join(project_folder, versioned_json_name)
        csv_path = os.path.join(project_folder, versioned_csv_name)

    # --- END of MODIFICATIONS ---
    
    # SSH connection info
    target = scan_data.get("target", "")
    port = scan_data.get("port", 22)
    ip = target

    ssh_info = {
        "username": scan_data.get("username"),
        "password": scan_data.get("password"),
        "ip": ip,
        "port": int(port) or 22,
    }

    try:
        if method not in ("remote", "agent"):
            print("[-] Unsupported audit method for Linux.")
            return None, None

        # Pass the final (potentially versioned) paths to the ubuntu function
        ubuntu(
            standard=standard,
            version=version,
            exclude_audits=excluded,
            method=method,
            ssh_info=ssh_info if method == "remote" else None,
            output_file=script_path,
            json_path=json_path,
            csv_path=csv_path
        )

        if method == "remote":
            return csv_path, json_path if os.path.exists(json_path) else None

        return script_path, None

    except Exception as e:
        print(f"[!] Linux audit failed: {e}")
        return None, None

def windows_config_audit(data):
    print("[*] Entered windows_config_audit()")
    scan_data = data.get("scan_data", {})
    base_dir = os.path.dirname(os.path.abspath(__file__))
    windows_dir = os.path.join(base_dir,"Configuration_Audit","Windows")
    
    # Get OS selection from scan_data 
    selected_os = (scan_data.get("complianceCategory"))
    audit_method = (scan_data.get("auditMethod")).strip().lower()

    # Normalize compliance and standard names
    compliance_name = (scan_data.get("complianceCategory") or "").strip().lower()
    print(compliance_name)
    standard = (scan_data.get("complianceSecurityStandard") or "").strip().lower()

    normalized_compliance = re.sub(r'[^\w._-]+', '', compliance_name)
    normalized_standard = re.sub(r'[^\w._-]+', '', compliance_name)
    print(f"[DEBUG] normalized_compliance: {normalized_compliance}")
    print(f"[DEBUG] normalized_standard: {normalized_standard}")
    project_name = (data.get("project_name") or "unknown_project").strip().replace(" ", "_")
    scan_name = (data.get("scan_name") or "unknown_scan").strip().replace(" ", "_")
    safe_compliance_name = normalized_compliance or "unknown_compliance"
    safe_standard = normalized_standard or "unknown_standard"
    safe_project_name = project_name
    safe_scan_name = scan_name
    base_dir = os.path.dirname(os.path.abspath(__file__))
    project_folder = os.path.join(
                                base_dir,
                                "Projects",
                                project_name,
                                scan_name
                            )
    # Create the nested directory structure if it doesn't exist
    os.makedirs(project_folder, exist_ok=True)
    
    base_dynamic_filename = f"{safe_compliance_name}_{safe_standard}_{safe_project_name}_{safe_scan_name}"
    dynamic_filename_extension = ".csv"
    dynamic_filename = f"{base_dynamic_filename}{dynamic_filename_extension}"
    result_csv = os.path.join(project_folder, dynamic_filename)
    version = 1
    while os.path.exists(result_csv):
            version += 1
            dynamic_filename = f"{base_dynamic_filename}_v{version}{dynamic_filename_extension}"
            result_csv = os.path.join(project_folder, dynamic_filename)

    try:
        # Dynamically find the query CSV
        queries_dir = os.path.join(base_dir,windows_dir,"CIS" , "Queries_Data")
        csv_name = (scan_data.get("complianceCategory") or "").strip() + ".csv"
        print(f"[DEBUG] Looking for CSV: {csv_name}")
        csv_path = os.path.join(queries_dir, csv_name)
        # Check if file exists
        if not os.path.exists(csv_path):
            print(f"[!] CSV file not found: {csv_path}")
            return None, None

        validate_dir = os.path.join(base_dir,windows_dir,"CIS", "Validate_Data")
        validate_csv_name = (scan_data.get("complianceCategory") or "").strip() + "_Validate.csv"
        print(f"[DEBUG] Looking for validation CSV: {validate_csv_name}")
        if not validate_csv_name:
            print("[!] Validation CSV name is empty.")
            return None, None
        validate_csv_path = os.path.join(validate_dir, validate_csv_name)
        
        output_json_remote_path = f"C:\\Windows\\Temp\\{os.path.basename(csv_path).replace('.csv', '_output.json')}"
        
        output_json_local_path = os.path.join(project_folder,f"{base_dynamic_filename}.json")

        script_file_name = f"{base_dynamic_filename}.ps1"

        generate_script_path = os.path.join(base_dir, windows_dir, "Output", script_file_name)

        print(f"[DEBUG] PowerShell script will be generated at: {generate_script_path}")

        remote_script_path = "C:\\Windows\\Temp\\generate_script.ps1"
        excluded_queries = scan_data.get("uncheckedComplianceItems", [])


        if audit_method == "remoteaccess":
            username = scan_data.get("username")
            password = scan_data.get("password")
            target_ip = scan_data.get("target")
            
            # Generate PowerShell Script
            generate_script(csv_path, generate_script_path, output_json_remote_path, excluded_queries,audit_method)
            print(f"[+] PowerShell script generated at: {generate_script_path}")
            
            final_json_path = run_remote_audit(username, password, target_ip, generate_script_path, output_json_local_path)
            print("[*] Remote audit completed.")

            secpol_file_path = "C:\\secpol.cfg"
            cleanup_remote_files(username, password, target_ip, remote_script_path, output_json_remote_path, secpol_file_path)

            json_base_name = os.path.splitext(os.path.basename(final_json_path))[0]
            output_validation_csv_path = os.path.join(base_dir,windows_dir, f"Output/{json_base_name}_validation_result.csv")

            validate_compliance(final_json_path, validate_csv_path, output_validation_csv_path)
            print("[*] Validation completed successfully.")

            return final_json_path, output_validation_csv_path
        
        elif audit_method == "agent":
            # Generate PowerShell Script for agent method
            generate_script(csv_path, generate_script_path, None, excluded_queries,audit_method)
            print(f"[+] PowerShell script generated at: {generate_script_path}")
            print("[*] Using agent method.")
            script_path = download_script(generate_script_path)
            print(f"[+] Script downloaded to: {script_path}")
            return script_path, output_json_local_path
        else:
            print("[*] Local agent mode is not yet implemented.")
            return None, None
        
         
        
    except Exception as e:
            print(f"[!] Exception during Windows audit: {e}")
            return None, None



def download_script(script_path, second_file_path=None):
    print("[*] Entered download_script()")

    # Case 1: Only one file to download
    if second_file_path is None:
        if not os.path.isfile(script_path):
            print(f"[-] Script file not found: {script_path}")
            return None
        print(f"[+] Script ready for direct download: {script_path}")
        return script_path

    # Case 2: ZIP both files
    if not os.path.isfile(script_path):
        print(f"[-] Primary script not found: {script_path}")
        return None

    if not os.path.isfile(second_file_path):
        print(f"[-] Second file not found: {second_file_path}")
        return None

    zip_filename = os.path.join(os.path.dirname(script_path), "scripts_bundle.zip")
    try:
        with zipfile.ZipFile(zip_filename, 'w') as zipf:
            zipf.write(script_path, arcname=os.path.basename(script_path))
            zipf.write(second_file_path, arcname=os.path.basename(second_file_path))
        print(f"[+] Both files compressed into: {zip_filename}")
        return zip_filename
    except Exception as e:
        print(f"[!] Failed to create ZIP file: {e}")
        return None

    




def convert_csv_to_excel(csv_file_path, excel_file_path=None):
    """
    Converts a CSV file to an Excel (.xlsx) file with:
    - Dark blue header row
    - Green background for 'pass' in Status column
    - Red background for 'fail' in Status column
    - Centered alignment for all cells with wrapped text
    - Auto-sized columns and row heights to show full content

    Args:
        csv_file_path (str): Path to the input CSV file.
        excel_file_path (str, optional): Path to save the output Excel file.

    Returns:
        str: Path to the created Excel file.
    """
    import os
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    if not os.path.exists(csv_file_path):
        raise FileNotFoundError(f"CSV file not found: {csv_file_path}")

    # Load CSV into a DataFrame
    with open(csv_file_path, "r", encoding="utf-8", errors="replace") as f:
         df = pd.read_csv(f)


    # Determine output file path
    if not excel_file_path:
        excel_file_path = os.path.splitext(csv_file_path)[0] + ".xlsx"

    # Write DataFrame to Excel (no formatting yet)
    df.to_excel(excel_file_path, index=False, engine='openpyxl')

    # Open the workbook and worksheet for formatting
    wb = load_workbook(excel_file_path)
    ws = wb.active

    # Define styles
    header_fill = PatternFill(start_color="FF0921D6", end_color="FF0921D6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    pass_fill = PatternFill(start_color="FF0DD609", end_color="FF0DD609", fill_type="solid")
    fail_fill = PatternFill(start_color="FFD60909", end_color="FFD60909", fill_type="solid")

    status_col_index = None

    # Format header row
    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

        if cell.value and str(cell.value).strip().lower() == "status":
            status_col_index = col_num

    # Format all data cells (centered alignment and wrap text)
    for row in ws.iter_rows(min_row=2):
        max_lines = 1
        for cell in row:
            cell.alignment = center_alignment
            if isinstance(cell.value, str):
                lines = cell.value.count('\n') + 1
                max_lines = max(max_lines, lines)
        ws.row_dimensions[row[0].row].height = max(15, 13.5 * max_lines)

    # Apply coloring to Status column values
    if status_col_index:
        for row in ws.iter_rows(min_row=2, min_col=status_col_index, max_col=status_col_index):
            for cell in row:
                value = str(cell.value).strip().lower()
                if value == "pass":
                    cell.fill = pass_fill
                elif value == "fail":
                    cell.fill = fail_fill

    # Auto-adjust column widths based on longest line in content
    col_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                col_letter = get_column_letter(cell.column)
                value = str(cell.value)
                estimated_width = max(len(line) for line in value.split('\n'))
                col_widths[col_letter] = max(col_widths.get(col_letter, 0), estimated_width)

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = min(width * 1.2, 100)  # scale width, cap at 100

    # Save the styled Excel file
    wb.save(excel_file_path)

    return excel_file_path





def windows_compromise_assesment(data):
    print("[*] Entered windows_compromise_assesment()")
    scan_data = data.get("scan_data", {})
    base_dir = os.path.dirname(os.path.abspath(__file__))
    windows_dir = os.path.join(base_dir,"Compromise_Assesment","Windows")
    audit_method = (scan_data.get("auditMethod")).lower()
    print(f"[DEBUG] Audit method: {audit_method}")
    
    try:
        # Dynamically find the query CSV
        command_dir = os.path.join(base_dir,windows_dir, "Command_Data")
        command_file = "command.xlsx"
        print(f"[DEBUG] Looking for CSV: {command_file}")
        xlsx_path = os.path.join(command_dir, command_file)
        # Check if file exists
        if not os.path.exists(xlsx_path):
            print(f"[!] CSV file not found: {xlsx_path}")
            return None, None

        validate_dir = os.path.join(base_dir,windows_dir,"Validate_Data")
        validate_py_name = "validate.py"
        print(f"[DEBUG] Looking for validation CSV: {validate_py_name}")
        if not validate_py_name:
            print("[!] Validation py name is empty.")
            return None, None
        validate_csv_path = os.path.join(validate_dir, validate_py_name)
        


        # output_json_remote_path = f"C:\\Windows\\Temp\\{os.path.basename(csv_path).replace('.csv', '_output.json')}"
        output_json_local_path = os.path.join(base_dir,windows_dir, "Output", os.path.basename(xlsx_path).replace('.xlsx', '_output.json'))

        script_file_name = f"Generate_Script_IOC.ps1"
        generate_script_path = os.path.join(base_dir, windows_dir, "Output", script_file_name)
        print(f"[DEBUG] PowerShell script will be generated at: {generate_script_path}")

        # remote_script_path = "C:\\Windows\\Temp\\generate_script.ps1"

        excluded_controls = scan_data.get("uncheckedComplianceItems", [])
        


        if audit_method == "remoteaccess":
            username = scan_data.get("username")
            password = scan_data.get("password")
            target_ip = scan_data.get("target")
            
            # Generate PowerShell Script
            # generate_powershell_script( xlsx_path, generate_script_path, validate_csv_path)
            # print(f"[+] PowerShell script generated at: {generate_script_path}")
            
            # final_json_path = run_remote_audit(username, password, target_ip, generate_script_path, output_json_local_path)
            # print("[*] Remote audit completed.")

            # secpol_file_path = "C:\\secpol.cfg"
            # cleanup_remote_files(username, password, target_ip, remote_script_path, output_json_remote_path, secpol_file_path)

            # json_base_name = os.path.splitext(os.path.basename(final_json_path))[0]
            # output_validation_csv_path = os.path.join(base_dir,windows_dir, f"Output/{json_base_name}_validation_result.csv")

            # validate_compliance(final_json_path, validate_csv_path, output_validation_csv_path)
            # print("[*] Validation completed successfully.")

            # return final_json_path, output_validation_csv_path
        
        elif audit_method == "agent":
            # Generate PowerShell Script for agent method
            generate_powershell_script(xlsx_path, generate_script_path, excluded_controls,audit_method)
            print(f"[+] PowerShell script generated at: {generate_script_path}")
            print("[*] Using agent method.")
            script_path = download_script(generate_script_path)
            print(f"[+] Script downloaded to: {script_path}")
            return script_path, output_json_local_path
        else:
            print("[*] Local agent mode is not yet implemented.")
            return None, None
        
         
        
    except Exception as e:
            print(f"[!] Exception during Windows audit: {e}")
            return None, None
    

@api_view(['GET'])
@permission_classes([AllowAny])
def get_csv_file(request, folder_path, filename):
    import os
    from django.http import HttpResponse
    print("get csv file")

    base_dir = os.path.dirname(os.path.abspath(__file__))
    safe_folder_path = os.path.normpath(folder_path).replace('..', '')
    csv_dir = os.path.join(base_dir, safe_folder_path)
    csv_path = os.path.join(csv_dir, f'{filename}.csv')

    if not os.path.exists(csv_path):
        return HttpResponse('File not found', status=404)

    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return HttpResponse(content, content_type='text/csv')
    except Exception as e:
        return HttpResponse(f'Error reading file: {str(e)}', status=500)
    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_project_excels(request, project_name):
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        project_dir = os.path.join(base_dir, "Projects", project_name)

        if not os.path.exists(project_dir):
            return Response({'error': f"Project '{project_name}' not found."}, status=status.HTTP_404_NOT_FOUND)

        zip_buffer = BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            found_excel = False
            for scan_folder in os.listdir(project_dir):
                scan_path = os.path.join(project_dir, scan_folder)
                if os.path.isdir(scan_path):
                    for file_name in os.listdir(scan_path):
                        if file_name.lower().endswith('.xlsx'):
                            found_excel = True
                            file_path = os.path.join(scan_path, file_name)
                            arcname = os.path.join(scan_folder, file_name)
                            zip_file.write(file_path, arcname)

        if not found_excel:
            return Response({'error': 'No Excel files found for this project.'}, status=status.HTTP_404_NOT_FOUND)

        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{project_name}_scans.zip"'
        return response

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_project_scan_excels(request, project_name, scan_name):
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        scan_dir = os.path.join(base_dir, "Projects", project_name, scan_name)

        if not os.path.exists(scan_dir):
            return Response({'error': f"Scan folder '{scan_name}' for project '{project_name}' not found."}, status=status.HTTP_404_NOT_FOUND)

        zip_buffer = BytesIO()
        found_excel = False

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for file_name in os.listdir(scan_dir):
                if file_name.lower().endswith(".xlsx"):
                    found_excel = True
                    file_path = os.path.join(scan_dir, file_name)
                    zip_file.write(file_path, arcname=file_name)

        if not found_excel:
            return Response({'error': f"No Excel files found in '{scan_name}'."}, status=status.HTTP_404_NOT_FOUND)

        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type="application/zip")
        response['Content-Disposition'] = f'attachment; filename="{project_name}_{scan_name}_excels.zip"'
        return response

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)