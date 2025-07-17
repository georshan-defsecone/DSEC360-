import csv
import json
import sys
import re
import os
import traceback
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

def extract_audit_id_from_json_name(json_name):
    parts = json_name.strip('_').split('_')
    id_parts = []
    for part in parts:
        if part.isdigit():
            id_parts.append(part)
        else:
            break
    return '.'.join(id_parts) if len(id_parts) >= 2 else None

def extract_audit_id_from_csv_name(audit_name):
    match = re.match(r'^(\d+(?:\.\d+)+)', audit_name.strip())
    return match.group(1) if match else None

def strip_audit_number(audit_name):
    return re.sub(r'^\d+(?:\.\d+)+\s*', '', audit_name).strip()

def load_csv(csv_file):
    expected_dict = {}
    with open(csv_file, newline='') as f:  
        reader = csv.DictReader(f)
        for row in reader:
            full_name = row['name'].strip()
            audit_id = extract_audit_id_from_csv_name(full_name)
            if not audit_id:
                continue
            expected_dict[audit_id] = {
                'audit_name': full_name,
                'expected_value': row['expected_value'].strip(),
                'comparison_type': row['comparison_type'].strip(),
                'field_to_check': row['field_to_check'].strip(),
                'description': row.get('description', '').strip(),
                'remediation': row.get('remediation', '').strip()
            }
    return expected_dict

def compare_values(actual, expected, comp_type):
    try:
        expected_is_null = expected.strip().lower() == 'null'
        actual_is_null = actual is None

        if comp_type == 'equals':
            return actual_is_null if expected_is_null else str(actual) == expected
        elif comp_type == 'in_list':
            return actual_is_null if expected_is_null else str(actual) in [val.strip() for val in expected.split(',')]
        elif comp_type == 'max_value':
            return False if actual_is_null or expected_is_null else float(actual) <= float(expected)
        elif comp_type == 'min_value':
            return False if actual_is_null or expected_is_null else float(actual) >= float(expected)
        elif comp_type == 'contains':
            return False if expected_is_null or actual_is_null else expected in str(actual)
        return False
    except Exception:
        return False

def write_csv(results, csv_path):
    fieldnames = ['CIS.NO', 'Subject', 'Description', 'Current Settings', 'Status', 'Remediation']
    with open(csv_path, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in results:
            writer.writerow(row)

def write_excel(results, excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Report"

    headers = ['CIS.NO', 'Subject', 'Description', 'Current Settings', 'Status', 'Remediation']
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bold_font = Font(bold=True)

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center")

    for row_data in results:
        row = [
            row_data['CIS.NO'],
            row_data['Subject'],
            row_data['Description'],
            row_data['Current Settings'],
            row_data['Status'],
            row_data['Remediation']
        ]
        ws.append(row)
        status_cell = ws.cell(row=ws.max_row, column=5)
        if row_data['Status'] == 'PASS':
            status_cell.fill = pass_fill
        elif row_data['Status'] == 'FAIL':
            status_cell.fill = fail_fill

    wb.save(excel_path)

def validate(json_file, expected_dict, output_path):
    with open(json_file) as f:
        try:
            json_objects = json.load(f)
        except json.JSONDecodeError as e:
            print(f"❌ Failed to parse JSON file: {e}")
            traceback.print_exc()
            return

    results_csv = []
    seen_audit_ids = set()

    for data in json_objects:
        json_name = data.get('name', '').strip()
        audit_id = extract_audit_id_from_json_name(json_name)

        if not audit_id:
            results_csv.append({
                'CIS.NO': '',
                'Subject': json_name,
                'Status': 'FAIL',
                'Current Settings': 'Could not extract audit ID',
                'Description': '',
                'Remediation': ''
            })
            continue

        seen_audit_ids.add(audit_id)
        expected_info = expected_dict.get(audit_id)
        remediation = expected_info.get('remediation', '') if expected_info else ''
        if not expected_info:
            results_csv.append({
                'CIS.NO': audit_id,
                'Subject': strip_audit_number(json_name),
                'Status': 'FAIL',
                'Current Settings': f"No expected value found for audit ID: {audit_id}",
                'Description': '',
                'Remediation': remediation
            })
            continue

        expected_value = expected_info['expected_value']
        comparison_type = expected_info['comparison_type']
        field_to_check = expected_info['field_to_check']
        audit_name = expected_info['audit_name']
        audit_clean = strip_audit_number(audit_name)
        description = expected_info.get('description', '')
        results = data.get('results', None)

        if comparison_type == 'some_result':
            status = 'PASS' if results and isinstance(results, list) and len(results) > 0 else 'FAIL'
            msg = 'Results found' if status == 'PASS' else 'No results found'
            results_csv.append({
                'CIS.NO': audit_id,
                'Subject': audit_clean,
                'Status': status,
                'Current Settings': msg,
                'Description': description,
                'Remediation': '' if status == 'PASS' else remediation
            })
            continue

        if field_to_check.lower() == 'null':
            expected_is_null = expected_value.strip().lower() == 'null'
            status, msg = '', ''
            if expected_is_null:
                if results is None or (isinstance(results, list) and len(results) == 0):
                    status, msg = 'PASS', 'Results are null or empty'
                else:
                    status, msg = 'FAIL', f"{results}"
            else:
                if results is None or (isinstance(results, list) and len(results) == 0):
                    status, msg = 'FAIL', 'Expected value present but got null'
                else:
                    status, msg = 'PASS', 'Results present as expected'

            results_csv.append({
                'CIS.NO': audit_id,
                'Subject': audit_clean,
                'Status': status,
                'Current Settings': msg,
                'Description': description,
                'Remediation': '' if status == 'PASS' else remediation
            })
            continue

        if not isinstance(results, list):
            msg = "'results' is null" if results is None else str(results)
            results_csv.append({
                'CIS.NO': audit_id,
                'Subject': audit_clean,
                'Status': 'FAIL',
                'Current Settings': msg,
                'Description': description,
                'Remediation': remediation
            })
            continue

        all_passed = True
        fail_reasons = []

        for result in results:
            actual_value = result.get(field_to_check)
            if actual_value is None and expected_value.strip().lower() != 'null':
                fail_reasons.append(f"Missing '{field_to_check}'")
                all_passed = False
                continue

            if not compare_values(actual_value, expected_value, comparison_type):
                fail_reasons.append(f"{field_to_check} = {actual_value}")
                all_passed = False

        results_csv.append({
            'CIS.NO': audit_id,
            'Subject': audit_clean,
            'Status': 'PASS' if all_passed else 'FAIL',
            'Current Settings': 'All checks passed' if all_passed else '; '.join(fail_reasons),
            'Description': description,
            'Remediation': '' if all_passed else remediation
        })

    for audit_id, info in expected_dict.items():
        if audit_id not in seen_audit_ids:
            results_csv.append({
                'CIS.NO': audit_id,
                'Subject': strip_audit_number(info['audit_name']),
                'Status': 'FAIL',
                'Current Settings': 'Audit ID not found in JSON',
                'Description': info.get('description', ''),
                'Remediation': info.get('remediation', '')
            })

    # Save CSV
    write_csv(results_csv, output_path)

    # Save Excel with same base name
    excel_path = os.path.splitext(output_path)[0] + '.xlsx'
    write_excel(results_csv, excel_path)

    print(f"\n✅ Validation reports saved:\n- CSV: {output_path}\n- Excel: {excel_path}")

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python validate.py <check.csv> <out.json> <result.csv>")
        sys.exit(1)

    csv_path = sys.argv[1]
    json_path = sys.argv[2]
    result_csv = sys.argv[3]

    expected_rules = load_csv(csv_path)
    validate(json_path, expected_rules, result_csv)
