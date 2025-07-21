import json
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

def load_tsv_result(tsv_path):
    results = []
    with open(tsv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            results.append({
                "audit_id": row.get("CIS.No.", "").strip(),
                "audit_name": row.get("Name", "").strip(),
                "status": row.get("Result", "").strip(),
                "output": row.get("Output", "").strip(),  # or custom if needed
                "description": "",  # will be filled in later
                "remediation": ""   # will be filled in later
            })
    return results

def write_excel(results, excel_path):
    """
    Writes a list of result dictionaries to a styled Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Report"

    # Define headers and cell styling
    headers = ["CIS.NO", "Subject", "Description", "Current Settings", "Status", "Remediation"]
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") # Light Blue
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # Light Green
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Light Red
    bold_font = Font(bold=True)

    # Write and style the header row
    ws.append(headers)
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write the data rows and apply conditional formatting
    for row_data in results:
        row_values = [
            row_data.get("CIS.NO", ""),
            row_data.get("Subject", ""),
            row_data.get("Description", ""),
            row_data.get("Current Settings", ""),
            row_data.get("Status", ""),
            row_data.get("Remediation", "")
        ]
        ws.append(row_values)
        
        # Apply color to the 'Status' cell based on its value (Column E)
        status_cell = ws.cell(row=ws.max_row, column=5)
        if row_data.get("Status") == "PASS":
            status_cell.fill = pass_fill
        elif row_data.get("Status") == "FAIL":
            status_cell.fill = fail_fill

    # Auto-adjust column widths for readability
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) if max_length < 50 else 50
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(excel_path)

def validateResult(json_path, csv_path, output_csv_path):
    
    if json_path.endswith(".tsv"):
        json_data = load_tsv_result(json_path)
    else:
        with open(json_path, "r", encoding="utf-8-sig") as jf:
            json_data = json.load(jf)
    # Load metadata from the original CSV
    csv_data = {}
    with open(csv_path, "r", encoding="utf-8-sig") as cf:
        reader = csv.DictReader(cf)
        for row in reader:
            row = {k.replace('\ufeff', ''): v for k, v in row.items()}  # in case BOM sneaks in
            audit_id = row["audit_id"].strip()
            csv_data[audit_id] = {
                "description": row["audit_description"].strip(),
                "remediation": row["audit_remediation"].strip()
            }

    # Enrich JSON with metadata
    for audit in json_data:
        audit_id = audit.get("audit_id")
        if audit_id in csv_data:
            audit["description"] = csv_data[audit_id]["description"]
            audit["remediation"] = csv_data[audit_id]["remediation"]
        else:
            audit["description"] = ""
            audit["remediation"] = ""

    # Prepare rows for CSV
    csv_rows = []
    for audit in json_data:
        status = audit.get("status", "").strip().upper()
        output = audit.get("output", "").strip()
        print(output)
        # Extract current settings based on ** PASS ** or ** FAIL **
        if "** FAIL **" in output:
            current_settings = output.split("** FAIL **", 1)[1].strip()
        elif "** PASS **" in output:
            current_settings = output.split("** PASS **", 1)[1].strip()
        else:
            current_settings = output  # fallback

        remediation = audit.get("remediation", "") if status == "FAIL" else ""

        csv_rows.append({
            "CIS.NO": audit.get("audit_id", ""),
            "Subject": audit.get("audit_name", ""),
            "Description": audit.get("description", ""),
            "Current Settings": current_settings,
            "Status": status,
            "Remediation": remediation
        })

    # Write to output CSV
    fieldnames = ["CIS.NO", "Subject", "Description", "Current Settings", "Status", "Remediation"]
    with open(output_csv_path, "w", encoding="utf-8", newline="") as out_csv:
        writer = csv.DictWriter(out_csv, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(csv_rows)

    print(f"CSV with enriched audit results saved as: {output_csv_path}")
    
    excel_path = os.path.splitext(output_csv_path)[0] + '.xlsx'
    write_excel(csv_rows, excel_path)
    print(f"Excel report saved as: {excel_path}")