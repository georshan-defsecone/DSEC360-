from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.decorators import api_view, permission_classes
from rest_framework import status
from rest_framework.response import Response
from .models import Project, Scan
from .serializers import ProjectSerializer, ScanSerializer, ScanVersionSerializer
#from rest_framework.permissions import IsAuthenticated
from rest_framework.permissions import AllowAny
from django.http import JsonResponse
import uuid
from io import BytesIO,StringIO
import os
from pathlib import Path
import csv
import io
import traceback
from django.conf import settings
import pandas as pd
import openpyxl
from rest_framework_simplejwt.views import TokenObtainPairView
from .serializers import MyTokenObtainPairSerializer
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny,IsAuthenticated
from rest_framework.response import Response
from .permissions import IsAdminUserCustom
from rest_framework import status
from rest_framework_simplejwt.views import TokenObtainPairView
from .serializers import MyTokenObtainPairSerializer
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt
from startScan.views import database_config_audit, windows_config_audit, linux_config_audit
from startScan.views import windows_compromise_assesment
from startScan.views import download_script
from startScan.views import convert_csv_to_excel
from django.http import FileResponse
import json
import mimetypes
from django.shortcuts import get_object_or_404
from startScan.Configuration_Audit.Database.ORACLE import oraclevalidate as oraclevalidator
from startScan.Configuration_Audit.Database.MARIA import validate as mariavalidator
from startScan.Configuration_Audit.Database.MSSQL import validate_result as mssqlvalidator
from startScan.Configuration_Audit.Linux import validate as linuxvalidator

def sanitize(name):
    return re.sub(r'\W+', '', name).lower()


@api_view(['POST'])
@permission_classes([AllowAny])
def add_MyProjectsView(request):
    serializer = ScanSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([AllowAny])
def update_MyProjectsView(request, scan_id):
    try:
        scan = Scan.objects.get(scan_id=scan_id)
    except Scan.DoesNotExist:
        return Response({"error": "Scan not found"}, status=status.HTTP_404_NOT_FOUND)

    trash_value = request.data.get('trash')
    if trash_value is None:
        return Response({"error": "Missing 'trash' value in request body"}, status=status.HTTP_400_BAD_REQUEST)

    scan.trash = trash_value
    scan.save()
    return Response({"message": f"Scan trash status updated to {scan.trash}"}, status=status.HTTP_200_OK)
    



# Create your views here.




class MyTokenObtainPairView(TokenObtainPairView):
    serializer_class = MyTokenObtainPairSerializer

@api_view(['GET'])
@permission_classes([AllowAny])
def get_compliance_data(request, os_name):
    try:
        file_path = os.path.join(settings.BASE_DIR, 'scan', 'data', 'Configuration_Audit - Copy.xlsx')

        if not os.path.exists(file_path):
            return Response({'error': 'File not found'}, status=404)

        df = pd.read_excel(file_path)
        print("DataFrame loaded successfully!")
        print(os_name)
        df = df[df['Configuration Name'].str.lower() == os_name.lower()]
        json_data = df.to_dict(orient='records')
        return Response(json_data, status=200)

    except Exception as e:
        print("Error while loading Excel:", e)
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_compromise_assessment_data(request, os_name):
    try:
        file_path = os.path.join(settings.BASE_DIR, 'scan', 'data', 'Compromise_Assessment.xlsx')

        if not os.path.exists(file_path):
            return Response({'error': 'File not found'}, status=404)

        df = pd.read_excel(file_path)
        print("DataFrame loaded successfully!")

        df = df[df['OS'].str.lower() == os_name.lower()]
        json_data = df.to_dict(orient='records')
        return Response(json_data, status=200)

    except Exception as e:
        print("Error while loading Excel:", e)
        return Response({'error': str(e)}, status=500)
    

#get projects
@api_view(['GET'])
@permission_classes([IsAuthenticated])  # Only authenticated users
def get_projects_view(request):
    user = request.user
    projects = Project.objects.filter(
        Q(project_author=user.username) | Q(scans__scan_author=user.username),
        trash=False
    ).distinct()
    serializer = ProjectSerializer(projects, many=True)
    return Response(serializer.data)


#get project by id
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_project_by_id(request, project_id):
    try:
        project = Project.objects.get(project_id=project_id)
        serializer = ProjectSerializer(project)
        return Response(serializer.data)
    except Project.DoesNotExist:
        return Response({'error': 'Project not found'}, status=status.HTTP_404_NOT_FOUND)
    
#get projects (for allprojects in frontend)
@api_view(['GET'])
@permission_classes([IsAdminUserCustom])
def get_all_projects_view(request):
    projects = Project.objects.filter(trash=False)  # Only non-trashed projects
    serializer = ProjectSerializer(projects, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])  # Only authenticated users can create a project
def create_project_view(request):
    # Ensure that the logged-in user is set as the project_author
    data = request.data.copy()
    data['project_author'] = request.user.id  # Or you can use request.user.username

    serializer = ProjectSerializer(data=data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


#trash project
@api_view(['PUT'])
@permission_classes([IsAuthenticated])  
def update_project_view(request, project_id):
    try:
        project = Project.objects.get(project_id=project_id)
    except Project.DoesNotExist:
        return Response({'error': 'Project not found'}, status=status.HTTP_404_NOT_FOUND)

    serializer = ProjectSerializer(project, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


#trashed projects
@api_view(['GET'])
@permission_classes([IsAuthenticated])  
def trashed_projects_view(request):
    user = request.user

    if user.is_admin:
        # If the user is an admin, fetch all trashed projects
        trashed_projects = Project.objects.filter(trash=True)
    else:
        trashed_projects = Project.objects.filter(trash=True, project_author=user.username)
    serializer = ProjectSerializer(trashed_projects, many=True)
    #to check 
    # print(f"Logged in user: {user.username}")
    # for p in Project.objects.filter(trash=True):
    #     print(f"{p.project_id}: {p.project_author}")

    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def trashed_scans_view(request):
    user = request.user

    if user.is_admin:
        trashed_scans = Scan.objects.filter(trash=True)
    else:
        trashed_scans = Scan.objects.filter(trash=True, scan_author=user.username)

    serializer = ScanSerializer(trashed_scans, many=True)
    return Response(serializer.data)


#remove (trashed) project and thier related scans from DB
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_project_view(request, project_id):
    try:
        project = Project.objects.get(project_id=project_id)
    except Project.DoesNotExist:
        return Response({'error': 'Project not found'}, status=status.HTTP_404_NOT_FOUND)

    # Check if the user is the author of the project or an admin
    if request.user.username != project.project_author and not request.user.is_admin:
        return Response({'error': 'You do not have permission to delete this project'}, status=status.HTTP_403_FORBIDDEN)

    # Delete related scans
    project.scans.all().delete()  # assumes related_name='scans' on the FK

    # Delete the project
    project.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)



#remove (trashed) all projects and thier related scans from DB
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_all_projects_view(request):
    try:
        trashed_projects = Project.objects.filter(trash=True)
        for project in trashed_projects:
            project.scans.all().delete()  # Delete all related scans
            project.delete()  # Delete the project itself
        return Response({"message": "All trashed projects and related scans deleted."}, status=200)
    except Exception as e:
        return Response({"error": str(e)}, status=400)

#get scans
@api_view(['GET'])
@permission_classes([AllowAny])
def get_scans_view(request):
    scans = Scan.objects.filter(trash=False)
    serializer = ScanSerializer(scans, many=True)
    return Response(serializer.data)

#create scan only in djangorestframework
@api_view(['POST'])
@permission_classes([AllowAny])
def create_scan_view(request):
    serializer = ScanSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#trash scan
@api_view(['PUT'])
@permission_classes([AllowAny])
def update_scan_view(request, pk):
    try:
        scan = Scan.objects.get(pk=pk)
    except Scan.DoesNotExist:
        return Response({'error': 'Scan not found'}, status=status.HTTP_404_NOT_FOUND)

    serializer = ScanSerializer(scan, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#get scans by project id
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_project_scans_view(request, project_id):
    user = request.user

    try:
        project = Project.objects.get(project_id=project_id, trash=False)
    except Project.DoesNotExist:
        return Response({'error': 'Project not found'}, status=status.HTTP_404_NOT_FOUND)

    if user.is_admin:
        # Admin: show all scans for the project
        scans = Scan.objects.filter(project_id=project_id, trash=False)
    else:
        # Normal user: show only their scans in the project
        scans = Scan.objects.filter(
            project_id=project_id,
            scan_author=user.username,
            trash=False
        )

    serializer = ScanSerializer(scans, many=True)
    return Response(serializer.data)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_scan_view(request, scan_id):
    try:
        scan = Scan.objects.get(scan_id=scan_id)
    except Scan.DoesNotExist:
        return Response({'error': 'Scan not found'}, status=status.HTTP_404_NOT_FOUND)

    # Check if the user is the author of the scan or an admin
    if request.user.username != scan.scan_author and not request.user.is_admin:
        return Response({'error': 'You do not have permission to delete this scan'}, status=status.HTTP_403_FORBIDDEN)

    # Delete the scan
    scan.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)

#get scans by user logged in for result page
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_user_scans_view(request):
    user = request.user
    scans = Scan.objects.filter(scan_author=user.username, trash=False)
    serializer = ScanSerializer(scans, many=True)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([AllowAny])
def post_scan_file(request):
    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate file extension
        allowed_extensions = {'.txt', '.pdf', '.doc', '.docx', '.xls', '.xlsx'}
        file_ext = os.path.splitext(file.name)[1].lower()
        if file_ext not in allowed_extensions:
            return Response({'error': 'Invalid file type'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Create a secure upload directory if it doesn't exist
        upload_dir = os.path.join(settings.MEDIA_ROOT, 'target')
        os.makedirs(upload_dir, exist_ok=True)
        
        # Generate safe filename
        safe_filename = f"{uuid.uuid4()}{file_ext}"
        file_path = os.path.join(upload_dir, safe_filename)
        
        try:
            # Save file safely
            with open(file_path, 'wb+') as destination:
                for chunk in file.chunks():
                    destination.write(chunk)
            return Response({
                'message': 'File uploaded successfully',
                'filename': safe_filename
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response({
                'error': f'Error saving file: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
    return Response({'error': 'Invalid request method'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_scan_result_view(request, project_name, scan_name):
    user = request.user

    try:
        project = Project.objects.get(project_name=project_name, trash=False)
    except Project.DoesNotExist:
        return Response({'error': 'Project not found'}, status=status.HTTP_404_NOT_FOUND)

    try:
        if user.is_admin:
            scan = Scan.objects.get(scan_name=scan_name, project=project, trash=False)
        else:
            scan = Scan.objects.get(scan_name=scan_name, project=project, scan_author=user.username, trash=False)
    except Scan.DoesNotExist:
        return Response({'error': 'Scan not found or you do not have permission to view it.'}, status=status.HTTP_404_NOT_FOUND)

    serializer = ScanSerializer(scan)
    
    # Return the entire serialized scan data, which includes parsed_scan_result
    # This aligns with the frontend's expectation of `response.data` having `parsed_scan_result`
    return Response(serializer.data, status=status.HTTP_200_OK)




@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_project_details(request, project_id):
    try:
        project = Project.objects.get(project_id=project_id)
    except Project.DoesNotExist:
        return Response({"detail": "Project not found."}, status=status.HTTP_404_NOT_FOUND)

    serializer = ProjectSerializer(project, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def update_scan_result(request, scan_id, audit_id):
    try:
        new_status = request.data.get("status", "").strip().upper()
        if not new_status:
            return Response({"error": "Status is required."}, status=status.HTTP_400_BAD_REQUEST)

        scan = get_object_or_404(Scan, scan_id=scan_id)

        # 1. Get the current latest version data (which is a list of dicts from JSONField)
        if not scan.scan_result or scan.scan_result_version == 0:
            return Response({"error": "No scan results available to update."}, status=status.HTTP_404_NOT_FOUND)

        latest_version_key = f'v{scan.scan_result_version}'
        parsed_result_current_version = scan.scan_result.get(latest_version_key)

        # Ensure the retrieved data is a list; handle old CSV string format if encountered
        if not isinstance(parsed_result_current_version, list):
            if isinstance(parsed_result_current_version, str):
                try:
                    csv_file = StringIO(parsed_result_current_version)
                    reader = csv.DictReader(csv_file)
                    parsed_result_current_version = list(reader)
                    print(f"[DEBUG] Converted old CSV string from DB for version {latest_version_key}")
                except Exception as e:
                    print(f"\n🔥 Internal Server Error (Parsing old CSV in PUT): {e}")
                    traceback.print_exc()
                    return Response({"error": "Failed to parse existing scan result CSV for update."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            else:
                print(f"\n🔥 Internal Server Error (Unexpected type in PUT): {type(parsed_result_current_version)}")
                return Response({"error": f"Unexpected format for current scan result in version {latest_version_key}."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # Now, `parsed_result_current_version` is guaranteed to be a list of dictionaries.
        updated_in_memory = False
        for entry in parsed_result_current_version:
            # Check for "CIS.NO" (from CSV) OR "name" (from new JSON format)
            identifier = entry.get("CIS.NO") or entry.get("name")
            if str(identifier) == str(audit_id):
                entry["Status"] = new_status # Update the status in the in-memory list
                updated_in_memory = True
                break

        if not updated_in_memory:
            return Response({"error": "Audit ID not found in the current scan result version."}, status=status.HTTP_404_NOT_FOUND)

        # 2. Update the JSONField in the model for the *same* version key
        # We are modifying the content of the existing latest version in place.
        scan.scan_result[latest_version_key] = parsed_result_current_version
        
        # Save the scan instance. No version increment as we are updating in place.
        scan.save(update_fields=['scan_result']) # Only save the scan_result JSONField
        print(f"[DEBUG] Scan {scan_id} - Version {latest_version_key} updated in DB.")

        # 3. Update the corresponding Excel and CSV files on disk to reflect the change
        # `sync_excel_with_db_csv` now receives the *parsed list of dicts* directly
        updated_excel_path = sync_excel_with_db_csv(scan, parsed_result_current_version)
        
        return Response({
            "message": "Audit status updated successfully.",
            "excel_updated_path": updated_excel_path,
            # Optionally, you might want to return the updated item or whole scan:
            # "updated_item": entry 
        })

    except Exception as e:
        print("\n🔥 Internal Server Error in update_scan_result:")
        traceback.print_exc()
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def update_excel_status(scan, audit_id: str, new_status: str) -> str:
    """
    Updates the Excel file entry matching audit_id with the new status.
    Returns the updated file path or raises errors if fails.
    """

    project_name = scan.project.project_name
    scan_name = scan.scan_name

    compliance_name = scan.scan_data.get("complianceCategory", "").strip().lower().replace(" ", "")
    compliance_category = scan.scan_data.get("complianceSecurityStandard", "").strip().lower()

    # Build relative to current file (views.py)
    base_dir = Path(__file__).resolve().parent.parent  # This gives you backend/DSEC/
    excel_file_path = base_dir / "startScan" / "Projects" / project_name / scan_name / f"{compliance_name}_{compliance_category}_{project_name}_{scan_name}.xlsx"

    if not excel_file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_file_path}")

    # Load workbook
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    if "CIS.NO" not in headers or "Status" not in headers:
        raise ValueError("Missing 'CIS.NO' or 'Status' in Excel headers.")

    cis_index = headers.index("CIS.NO")
    status_index = headers.index("Status")

    found = False
    for row in ws.iter_rows(min_row=2):
        if str(row[cis_index].value).strip() == str(audit_id):
            row[status_index].value = new_status
            found = True
            break

    if not found:
        raise ValueError("Audit ID not found in Excel file.")

    wb.save(excel_file_path)

    return str(excel_file_path)


def sync_excel_with_db_csv(scan: Scan, parsed_scan_data: list[dict]) -> str:
    """
    Synchronizes the Excel file and its corresponding CSV on disk
    with the provided parsed scan data (list of dicts) for a specific scan version.
    Returns the Excel file path.
    """
    project_name = scan.project.project_name
    scan_name = scan.scan_name

    # Determine filename with version for both CSV and Excel
    # The version number comes from scan.scan_result_version (the latest version in DB)
    version_num = scan.scan_result_version
    
    compliance_name = scan.scan_data.get("complianceCategory", "").strip().lower().replace(" ", "")
    compliance_standard = scan.scan_data.get("complianceSecurityStandard", "").strip().lower()

    # Construct file paths
    base_dir = Path(__file__).resolve().parent.parent 
    project_dir = base_dir / "startScan" / "Projects" / project_name / scan_name

    csv_file_name = f"{compliance_name}_{compliance_standard}_{project_name}_{scan_name}_v{version_num}.csv"
    excel_file_name = f"{compliance_name}_{compliance_standard}_{project_name}_{scan_name}_v{version_num}.xlsx"

    csv_file_path = project_dir / csv_file_name
    excel_file_path = project_dir / excel_file_name

    # ✅ Ensure folder exists
    project_dir.mkdir(parents=True, exist_ok=True)

    # Convert the parsed_scan_data (list of dicts) back to CSV string
    if not parsed_scan_data: # Handle empty data case
        csv_text_to_write = ""
        fieldnames = ["CIS.NO", "Subject", "Description", "Current Settings", "Status", "Remediation"] # Default headers
    else:
        # Assuming all dicts in the list have the same keys, take from the first one
        fieldnames = list(parsed_scan_data[0].keys())
        output_csv = StringIO()
        writer = csv.DictWriter(output_csv, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(parsed_scan_data)
        csv_text_to_write = output_csv.getvalue()

    # ✅ Write the generated CSV string to the .csv file on disk
    with open(csv_file_path, "w", encoding="utf-8") as f:
        f.write(csv_text_to_write)
    print(f"[DEBUG] CSV file updated on disk: {csv_file_path}")

    # ✅ Rebuild (or update) the Excel file from the CSV file
    # This ensures consistency between CSV and Excel.
    convert_csv_to_excel(str(csv_file_path), str(excel_file_path)) # Uses the helper `convert_csv_to_excel`

    return str(excel_file_path) # Return path to the updated Excel


# --- update_excel_status (DEPRECATED or REMOVED/REFACTORED) ---
# This function is no longer needed as a standalone entry point.
# Its functionality is now integrated into `sync_excel_with_db_csv`.
# If you still have calls to this, you should replace them with calls to sync_excel_with_db_csv.
# For clarity, you should remove this function from your views.py entirely if it's unused.

@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_scan_output(request, scan_id):
    print(f"\n[DEBUG] Entered upload_scan_output for scan_id: {scan_id}")
    try:
        scan = get_object_or_404(Scan, scan_id=scan_id)
        project = scan.project

        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return JsonResponse({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        base_dir = Path(__file__).resolve().parent.parent 
        target_dir = base_dir / 'startScan' / 'Projects' / project.project_name / scan.scan_name
        os.makedirs(target_dir, exist_ok=True)

        file_path = target_dir / uploaded_file.name 
        with open(file_path, 'wb+') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)
        print(f"[DEBUG] Uploaded file saved to: {file_path}")

        if file_path.stat().st_size == 0:
            return JsonResponse({"error": "Uploaded file is empty."}, status=status.HTTP_400_BAD_REQUEST)

        scan_data = scan.scan_data or {}
        print(scan_data)
        flavour = scan_data.get('flavour', '')
        compliance_standard = scan_data.get('complianceSecurityStandard', '')
        compliance_category = scan_data.get('complianceCategory','')

        safe_flavour = sanitize(flavour)
        category = scan_data.get('category', '').lower()  # Ensure category is lowercase
        safe_standard = sanitize(compliance_standard)
        safe_project_name = sanitize(project.project_name)
        safe_scan_name = sanitize(scan.scan_name)
        safe_compliance_category = sanitize(compliance_category)
        print(safe_compliance_category) # Debug print from original code

        # Original lines for dynamic_filename and output_report_path
        # Keeping these as per your instruction "don't change anything else"
        # Note: This means output_report_path will use the version from when the scan was *created*
        # if this is not the first upload. For consistent versioned filenames,
        # these would ideally be calculated *after* `new_scan_result_version` (which you have
        # in create_scan, but not this `upload_scan_output` flow).
        # However, to strictly adhere to "don't change anything else", I'm leaving this as is.
        dynamic_filename = f"{safe_flavour}_{safe_standard}_{safe_project_name}_{safe_scan_name}.csv"
        output_report_path = target_dir / dynamic_filename

        if safe_flavour == 'oracle':
            csv_file_path = base_dir / 'startScan' / 'Configuration_Audit' / 'Database' / 'ORACLE' / 'CIS' / 'Validators' / 'check.csv'
            if not csv_file_path.exists():
                return JsonResponse({"error": f"CSV file not found at: {csv_file_path}"}, status=status.HTTP_400_BAD_REQUEST)

            expected_dict = oraclevalidator.load_csv(csv_file_path)
            oraclevalidator.validate(file_path, expected_dict, output_report_path)

        elif safe_flavour == 'maria':
            if safe_compliance_category == "mariadb10_11":
                csv_file_path = base_dir / 'startScan' / 'Configuration_Audit' / 'Database' / 'MARIA' / 'CIS' / 'Validators' / 'MariaDB_10_11_validate.csv'
            elif safe_compliance_category == "mariadb10_6":
                csv_file_path = base_dir / 'startScan' / 'Configuration_Audit' / 'Database' / 'MARIA' / 'CIS' / 'Validators' / 'MariaDB_10_6_validate.csv'
            else:
                return JsonResponse({"error": f"MariaDB compliance category '{compliance_category}' not supported for validation."}, status=status.HTTP_400_BAD_REQUEST)
            
            if not csv_file_path.exists():
                return JsonResponse({"error": f"CSV file not found at: {csv_file_path}"}, status=status.HTTP_400_BAD_REQUEST)

            mariavalidator.validate_maria_db(file_path, csv_file_path , output_report_path)

        elif safe_flavour == 'mssql':
            if safe_compliance_category == "microsoftsqlserver2019":
                csv_file_path = base_dir / 'startScan' / 'Configuration_Audit' / 'Database' / 'MSSQL' / 'CIS' / 'Result_Validators' / 'microsoft_sql_server_2019_validator.csv'
            elif safe_compliance_category == "microsoftsqlserver2016":
                csv_file_path = base_dir / 'startScan' / 'Configuration_Audit' / 'Database' / 'MSSQL' / 'CIS' / 'Result_Validators' / 'microsoft_sql_server_2016_validator.csv'
            elif safe_compliance_category == "microsoftsqlserver2017":
                csv_file_path = base_dir / 'startScan' / 'Configuration_Audit' / 'Database' / 'MSSQL' / 'CIS' / 'Result_Validators' / 'microsoft_sql_server_2017_validator.csv'
            elif safe_compliance_category == "microsoftsqlserver2022":
                csv_file_path = base_dir / 'startScan' / 'Configuration_Audit' / 'Database' / 'MSSQL' / 'CIS' / 'Result_Validators' / 'microsoft_sql_server_2022_validator.csv'
            else:
                return JsonResponse({"error": f"MSSQL compliance category '{compliance_category}' not supported for validation."}, status=status.HTTP_400_BAD_REQUEST)
            
            if not csv_file_path.exists():
                return JsonResponse({"error": f"CSV file not found at: {csv_file_path}"}, status=status.HTTP_400_BAD_REQUEST)

            mssqlvalidator.validate_mssql(file_path, csv_file_path , output_report_path)

        elif category == 'linux':
            # Construct the path to the Linux metadata CSV file
            # This path is based on the pattern from your generate.py script
            compliance_category = compliance_category.lower().replace(" ", "_")  # Ensure it's lowercase and underscores
            csv_metadata_path = base_dir / 'startScan' / 'Configuration_Audit' / 'Linux' / compliance_standard / (compliance_category + '.csv')
            
            if not csv_metadata_path.exists():
                return JsonResponse({"error": f"Linux metadata CSV not found at: {csv_metadata_path}"}, status=status.HTTP_400_BAD_REQUEST)
            
            # Call the Linux validator function
            # It takes the uploaded file (json/tsv), the metadata file, and the output path
            linuxvalidator.validateResult(json_path=str(file_path), csv_path=str(csv_metadata_path), output_csv_path=str(output_report_path))
        
        else:
            return JsonResponse({"error": f"Validation for flavour '{flavour}' is not implemented."}, status=status.HTTP_400_BAD_REQUEST)

        # --- START OF MODIFIED BLOCK ---
        # This part now correctly parses the CSV and saves it using update_scan_result
        # The `output_report_path` is the path to the validated CSV file generated by the validators.
        with open(output_report_path, 'r', encoding='utf-8', errors="replace") as f:
            csv_text_from_report = f.read() # Read the CSV content as a string
        
        # Parse the CSV text into a list of dictionaries
        try:
            csv_file_in_memory = StringIO(csv_text_from_report)
            reader = csv.DictReader(csv_file_in_memory)
            parsed_scan_result_list = list(reader) # This is the list of dicts we want to store
            print(f"[DEBUG] Parsed {len(parsed_scan_result_list)} rows from report CSV for storage.")
        except Exception as e:
            print(f"[!] Failed to parse generated report CSV for storage: {e}")
            traceback.print_exc()
            scan.scan_status = "failed_report_parse_upload" # Specific status for this failure
            scan.save(update_fields=['scan_status'])
            return JsonResponse({"error": f"Failed to parse generated validation report CSV for storage: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Use the model's method to store the parsed list as a new version
        scan.update_scan_result(parsed_scan_result_list) 
        
        # IMPORTANT: Refresh the `scan` object from the database to get its latest state
        # (especially the updated scan_result_version and the JSONField content)
        scan = Scan.objects.get(pk=scan.pk) 
        print(f"[DEBUG] Scan instance refreshed after update_scan_result. New version: {scan.scan_result_version}")

        # Synchronize Excel and CSV files on disk using the newly updated scan object and parsed data.
        # This will save the Excel/CSV with the correct versioned filename.
        sync_excel_with_db_csv(scan, parsed_scan_result_list) 

        # Update final scan status
        scan.scan_status = "complete"
        scan.save(update_fields=['scan_status'])
        # --- END OF MODIFIED BLOCK ---

    except Exception as e:
        print("\n🔥 Internal Server Error in upload_scan_output (main try-except):")
        traceback.print_exc()
        return JsonResponse({"error": f"Processing failed: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    return JsonResponse({
        "message": "File uploaded, validated, and scan result updated successfully.",
        "uploaded_file": str(file_path),
        "validation_report_csv": str(output_report_path), # This path is from the original calculation
        "validation_report_xlsx": str(output_report_path.with_suffix('.xlsx')), # This path is from the original calculation
        "new_scan_version": scan.scan_result_version # This will be the new version
    }, status=status.HTTP_200_OK)





@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def create_scan(request):
    print("entered create scan")
    request_data = request.data.copy()
    print(request_data)

    project_name = request_data.get("project_name")
    scan_author = request_data.get("scan_author", "unknown")
    scan_name = request_data.get("scan_name")
    scan_type = request_data.get("scan_type", "unknown_type")
    
    if not project_name:
        return Response({"error": "project_name is required."}, status=status.HTTP_400_BAD_REQUEST)

    if not scan_name:
        return Response({"error": "scanName is required."}, status=status.HTTP_400_BAD_REQUEST)

    project, created = Project.objects.get_or_create(
        project_name=project_name,
        defaults={
            "project_id": str(uuid.uuid4()),
            "project_author": scan_author,
            "trash": False,
        }
    )

    scan_instance = None
    is_new_scan = False

    try:
        scan_instance = Scan.objects.get(scan_name=scan_name, project=project, trash=False)
        print(f"Scan with name '{scan_name}' found. Preparing to execute and add a new result version.")
    except Scan.DoesNotExist:
        is_new_scan = True
        new_scan_payload = {
            "scan_id": str(uuid.uuid4()),
            "scan_name": scan_name,
            "scan_author": scan_author,
            "scan_type": scan_type,
            "scan_status": "Pending",
            "project": project.pk,
            "scan_data": request_data.get("scan_data", {})
        }
        serializer = ScanSerializer(data=new_scan_payload)
        if serializer.is_valid():
            scan_instance = serializer.save()
            print(f"Creating new scan with ID: {scan_instance.scan_id} (initial version v1 expected).")
        else:
            print("[!] Serializer Errors for new scan creation:", serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    if scan_instance is None:
        return Response({"error": "Failed to create or retrieve scan instance."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    scan_instance.scan_status = "Running"
    scan_instance.save(update_fields=['scan_status'])

    file_path, output_json_path = None, None
    try:
        # Pass the full request_data to launch_scan, as it contains all necessary info
        # This will return (result_csv_path, json_output_path) for remoteAccess
        # or (script_path, None) for agent method.
        # This `request_data` should contain `scan_id` from the newly created instance if `is_new_scan` is True.
        # Let's ensure scan_id is in request_data for launch_scan
        if is_new_scan:
            request_data['scan_id'] = scan_instance.scan_id

        file_path, output_json_path = launch_scan(request_data) 
        if file_path is None: 
            print("[!] launch_scan() returned None for file_path.")
            scan_instance.scan_status = "failed"
            scan_instance.save(update_fields=['scan_status'])
            return Response({"error": "Scan execution failed to produce a primary file."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        print(f"[DEBUG] Returned file_path: {file_path}")
        print(f"[DEBUG] file exists: {os.path.exists(file_path)}")
    except Exception as e:
        print(f"[!] Exception during launch_scan call: {e}")
        traceback.print_exc()
        scan_instance.scan_status = "failed"
        scan_instance.save(update_fields=['scan_status'])
        return Response({"error": f"Scan execution failed: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # Handle scan_output (JSON) from output_json_path
    if output_json_path and os.path.exists(output_json_path):
        with open(output_json_path, "r", encoding="utf-8") as f:
            try:
                output_data = json.load(f)
                scan_instance.scan_output = output_data
            except json.JSONDecodeError:
                print("[!] Failed to decode output.json")
    
    scan_result_content_to_store = None 
    response_file_to_return = None 

    audit_method = request_data.get("scan_data", {}).get("auditMethod", "").strip().lower()
    
    if audit_method == "agent":
        print("Processing agent script file for download.")
        filename_for_download = os.path.basename(file_path)
        mime_type, _ = mimetypes.guess_type(file_path)
        mime_type = mime_type or "application/octet-stream"
        response_file_to_return = FileResponse(
            open(file_path, "rb"),
            as_attachment=True,
            filename=filename_for_download,
            content_type=mime_type
        )
        scan_result_content_to_store = None 
        print("[DEBUG] Agent script prepared for download.")

    elif str(file_path).lower().endswith('.csv'): # Check explicitly for .csv
        print("Processing scan result (CSV from remoteAccess).")
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            csv_text = f.read()
        
        try:
            csv_file_in_memory = StringIO(csv_text)
            reader = csv.DictReader(csv_file_in_memory)
            scan_result_content_to_store = list(reader)
        except Exception as e:
            print(f"[!] Failed to parse CSV from {file_path}: {e}")
            traceback.print_exc()
            scan_instance.scan_status = "failed_csv_parse"
            scan_instance.save(update_fields=['scan_status'])
            return Response({"error": "Failed to parse scan result as CSV."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        scan_instance.update_scan_result(scan_result_content_to_store) 
        scan_instance = Scan.objects.get(pk=scan_instance.pk) # Refresh from DB!
        print(f"[DEBUG_VIEW] Refreshed scan_instance. New version: {scan_instance.scan_result_version}")
        
    else: # Fallback for unexpected file types or non-agent, non-CSV outputs
        print(f"[!] Unknown file type or handling for: {file_path} with audit method {audit_method}")
        scan_instance.scan_status = "failed_unknown_file_type"
        scan_instance.save(update_fields=['scan_status'])
        return Response({"error": f"Unknown file type or handling for {os.path.basename(file_path)}."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
    # Final status update.
    scan_instance.scan_status = "complete" if scan_instance.scan_status == "Running" else scan_instance.scan_status
    scan_instance.save(update_fields=['scan_status', 'scan_output'])

    if response_file_to_return:
        return response_file_to_return
    else:
        return Response({
            "message": "Scan initiated and processed successfully.",
            "scan": ScanSerializer(scan_instance).data
        }, status=status.HTTP_201_CREATED if is_new_scan else status.HTTP_200_OK)



import re

def launch_scan(data):
    print("[*] Entered launch_scan()")
    scan_data = data.get("scan_data", {})
    

    # Normalize and extract values from scan_data
    scan_type = (scan_data.get("scanType") or "").strip().lower().replace(" ", "").replace("_", "")
    print("[DEBUG] scan_type:", scan_type)

    os_name = (scan_data.get("os") or "").strip().lower()
    print("[DEBUG] os_name:", os_name)

    auth_type = (scan_data.get("auditMethod") or "").strip().lower()
    category=(scan_data.get("category")or"").strip().lower()
    compliance_name=(scan_data.get("complianceCategory") or "").strip().lower()

    standard = scan_data.get("complianceSecurityStandard")
    # Remove all non-alphanumeric characters including underscores, then lowercase
    normalized_compliance = re.sub(r'[\W_]+', '', compliance_name or "").lower()
    print(scan_data)
    print("[DEBUG] auth_type:", auth_type)
    category = (scan_data.get("category") or "").strip().lower()
    print("[DEBUG] category:", category)

    audit_method = (scan_data.get("auditMethod") or "").strip().lower()


    compliance_name = (scan_data.get("complianceCategory") or "").strip().lower()
    standard = scan_data.get("complianceSecurityStandard")

    # Remove all non-alphanumeric characters including underscores, then lowercase
    normalized_compliance = re.sub(r'[\W_]+', '', compliance_name or "").lower()
    print("[DEBUG] scan_data received:", scan_data)

    if scan_type == "configurationaudit":

        if category=="database":
           return database_config_audit(data)
           

        if category=="windows":
            
            return windows_config_audit(data)

        if category == "linux":
            return linux_config_audit(data)

        if category=="firewall":
            return None,None
        
        if category == "firewall":
            return None, None

    elif scan_type == "compromiseassessment":
        if category == "windows":
            return windows_compromise_assesment(data)

    
        if category == "linux":
            return None, None

           
    return None, None  